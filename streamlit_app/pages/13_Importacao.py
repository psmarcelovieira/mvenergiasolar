import io
from datetime import date

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import api_client

st.set_page_config(page_title="Importação em Lote", layout="wide")
from auth import verificar_auth, rodape_usuario
verificar_auth()
rodape_usuario()

st.title("📥 Importação em Lote")
st.caption("Baixe o template, preencha no Excel e faça upload para importar vários registros de uma vez.")

# ── HELPERS ───────────────────────────────────────────────────────────────────

def _template(colunas: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    hf = Font(bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor="E6B400")
    df = Font(italic=True, color="888888")
    db = PatternFill("solid", fgColor="F5F5F5")

    for i, col in enumerate(colunas, 1):
        c = ws.cell(row=1, column=i, value=col["header"])
        c.font, c.fill = hf, hb
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = col.get("w", 18)

        hint = col.get("hint", "")
        if col.get("req"):
            hint = f"* {hint}" if hint else "*"
        h = ws.cell(row=2, column=i, value=hint)
        h.font, h.fill = df, db

        if col.get("vals"):
            formula = '"' + ",".join(col["vals"]) + '"'
            dv = DataValidation(type="list", formula1=formula,
                                allow_blank=True, showDropDown=False)
            dv.sqref = f"{get_column_letter(i)}3:{get_column_letter(i)}1000"
            ws.add_data_validation(dv)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _str(val) -> str | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    return s or None


def _float(val) -> float | None:
    try:
        return float(val) if val not in (None, "", "nan") else None
    except Exception:
        return None


def _limpar(df: pd.DataFrame) -> pd.DataFrame:
    """Remove linha de dica se presente (começa com * ou Ex:)."""
    if df.empty:
        return df
    primeiro = str(df.iloc[0, 0]).strip()
    if primeiro.startswith(("*", "Ex:", "Exemplo")):
        df = df.iloc[1:].reset_index(drop=True)
    return df.fillna("")


def _executar(df: pd.DataFrame, fn, entidade: str):
    total = len(df)
    ok, erros = 0, []
    bar = st.progress(0, text="Importando…")
    for i, (_, row) in enumerate(df.iterrows()):
        bar.progress((i + 1) / total, text=f"Linha {i + 2}/{total + 1}")
        err = fn(row)
        if err:
            erros.append(f"Linha {i + 2}: {err}")
        else:
            ok += 1
    bar.empty()
    if ok:
        st.success(f"✅ {ok} {entidade} importado(s) com sucesso.")
    if erros:
        st.error(f"❌ {len(erros)} erro(s):")
        for e in erros:
            st.caption(e)


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

tab_cli, tab_prod, tab_colab, tab_proj = st.tabs([
    "👥 Clientes", "📦 Produtos", "👷 Colaboradores", "📋 Projetos",
])

# ── CLIENTES ──────────────────────────────────────────────────────────────────
with tab_cli:
    st.subheader("Importar Clientes")

    COLS_CLI = [
        {"header": "nome",          "hint": "João Silva",       "req": True,  "w": 26},
        {"header": "cpf_cnpj",      "hint": "12345678901",                    "w": 16},
        {"header": "telefone",      "hint": "85999990000",                    "w": 15},
        {"header": "email",         "hint": "joao@email.com",                 "w": 24},
        {"header": "cidade",        "hint": "Fortaleza",                      "w": 16},
        {"header": "uf",            "hint": "CE",                             "w":  6},
        {"header": "tipo_pessoa",   "hint": "PF",               "w":  8,
         "vals": ["PF", "PJ"]},
        {"header": "status_cliente","hint": "Prospecto",        "w": 13,
         "vals": ["Prospecto", "Ativo", "Inativo", "Bloqueado"]},
        {"header": "origem_lead",   "hint": "Indicação",        "w": 13,
         "vals": ["Indicação", "Google", "Instagram", "Facebook",
                  "WhatsApp", "Site", "Evento", "Outros"]},
        {"header": "tipo_imovel",   "hint": "Residencial",      "w": 14,
         "vals": ["Residencial", "Comercial", "Rural", "Industrial"]},
        {"header": "conta_energia", "hint": "350.00",                         "w": 14},
        {"header": "renda_mensal",  "hint": "5000.00",                        "w": 14},
        {"header": "profissao",     "hint": "Autônomo",                       "w": 16},
        {"header": "escolaridade",  "hint": "Superior",         "w": 14,
         "vals": ["Fundamental", "Médio", "Superior", "Pós-graduação"]},
        {"header": "endereco",      "hint": "Rua A, 100",                     "w": 24},
        {"header": "bairro",        "hint": "Centro",                         "w": 16},
    ]

    c1, _ = st.columns([1, 3])
    c1.download_button("⬇️ Template Clientes", _template(COLS_CLI),
                       "template_clientes.xlsx", XLSX_MIME)

    up = st.file_uploader("Excel preenchido", type="xlsx", key="up_cli")
    if up:
        df_cli = _limpar(pd.read_excel(up, dtype=str))
        st.dataframe(df_cli, use_container_width=True, height=180)
        st.caption(f"{len(df_cli)} linha(s).")

        if st.button("📥 Importar Clientes", type="primary", key="btn_cli"):
            def _fn_cli(row):
                nome = _str(row.get("nome"))
                if not nome:
                    return "campo 'nome' obrigatório"
                payload = {
                    "nome":         nome,
                    "cpf_cnpj":     _str(row.get("cpf_cnpj")),
                    "telefone":     _str(row.get("telefone")),
                    "email":        _str(row.get("email")),
                    "cidade":       _str(row.get("cidade")),
                    "uf":           (_str(row.get("uf")) or "")[:2].upper() or None,
                    "tipo_pessoa":  _str(row.get("tipo_pessoa")) or "PF",
                    "origem_lead":  _str(row.get("origem_lead")),
                    "tipo_imovel":  _str(row.get("tipo_imovel")),
                    "conta_energia":_float(row.get("conta_energia")),
                    "renda_mensal": _float(row.get("renda_mensal")),
                    "profissao":    _str(row.get("profissao")),
                    "escolaridade": _str(row.get("escolaridade")),
                    "endereco":     _str(row.get("endereco")),
                    "bairro":       _str(row.get("bairro")),
                }
                resp = api_client.criar_cliente(payload)
                if resp.status_code not in (200, 201):
                    return resp.json().get("detail", resp.text)
                return None
            _executar(df_cli, _fn_cli, "cliente(s)")

# ── PRODUTOS ──────────────────────────────────────────────────────────────────
with tab_prod:
    st.subheader("Importar Produtos")

    COLS_PROD = [
        {"header": "codigo",         "hint": "PAIN-001",      "req": True,  "w": 12},
        {"header": "nome",           "hint": "Painel 550W",   "req": True,  "w": 26},
        {"header": "categoria",      "hint": "Painel",        "req": True,  "w": 14,
         "vals": ["Painel", "Inversor", "Cabo", "Estrutura", "Acessório", "Serviço"]},
        {"header": "unidade",        "hint": "UN",            "req": True,  "w":  8,
         "vals": ["UN", "M", "KG", "KWP"]},
        {"header": "custo_unitario", "hint": "850.00",        "req": True,  "w": 14},
        {"header": "preco_venda",    "hint": "1250.00",       "req": True,  "w": 14},
        {"header": "fabricante",     "hint": "Risen",                       "w": 14},
        {"header": "modelo",         "hint": "RSM110-8-550M",               "w": 18},
        {"header": "potencia_w",     "hint": "550",                         "w": 12},
        {"header": "qtd_minima",     "hint": "0",                           "w": 12},
    ]

    c1, _ = st.columns([1, 3])
    c1.download_button("⬇️ Template Produtos", _template(COLS_PROD),
                       "template_produtos.xlsx", XLSX_MIME)

    up = st.file_uploader("Excel preenchido", type="xlsx", key="up_prod")
    if up:
        df_prod = _limpar(pd.read_excel(up, dtype=str))
        st.dataframe(df_prod, use_container_width=True, height=180)
        st.caption(f"{len(df_prod)} linha(s).")

        if st.button("📥 Importar Produtos", type="primary", key="btn_prod"):
            def _fn_prod(row):
                for campo in ("codigo", "nome", "categoria", "unidade",
                              "custo_unitario", "preco_venda"):
                    if not _str(row.get(campo)):
                        return f"campo '{campo}' obrigatório"
                payload = {
                    "codigo":         _str(row["codigo"]),
                    "nome":           _str(row["nome"]),
                    "categoria":      _str(row["categoria"]),
                    "unidade":        _str(row["unidade"]),
                    "custo_unitario": _float(row["custo_unitario"]),
                    "preco_venda":    _float(row["preco_venda"]),
                    "fabricante":     _str(row.get("fabricante")),
                    "modelo":         _str(row.get("modelo")),
                    "potencia_w":     _float(row.get("potencia_w")),
                    "qtd_minima":     int(_float(row.get("qtd_minima")) or 0),
                }
                resp = api_client.criar_produto(payload)
                if resp.status_code not in (200, 201):
                    return resp.json().get("detail", resp.text)
                return None
            _executar(df_prod, _fn_prod, "produto(s)")

# ── COLABORADORES ─────────────────────────────────────────────────────────────
with tab_colab:
    st.subheader("Importar Colaboradores")

    COLS_COLAB = [
        {"header": "nome",                "hint": "Maria Santos",  "req": True,  "w": 26},
        {"header": "cpf",                 "hint": "12345678901",                 "w": 14},
        {"header": "cargo",               "hint": "Vendedor",                    "w": 14},
        {"header": "tipo_contrato",       "hint": "Autônomo",      "req": True,  "w": 13,
         "vals": ["CLT", "Autônomo", "PJ"]},
        {"header": "percentual_comissao", "hint": "3  (= 3%)",     "req": True,  "w": 18},
        {"header": "telefone",            "hint": "85999990000",                 "w": 15},
        {"header": "email",               "hint": "maria@email.com",             "w": 24},
        {"header": "data_admissao",       "hint": "2026-05-26",    "req": True,  "w": 14},
    ]

    c1, _ = st.columns([1, 3])
    c1.download_button("⬇️ Template Colaboradores", _template(COLS_COLAB),
                       "template_colaboradores.xlsx", XLSX_MIME)

    up = st.file_uploader("Excel preenchido", type="xlsx", key="up_colab")
    if up:
        df_colab = _limpar(pd.read_excel(up, dtype=str))
        st.dataframe(df_colab, use_container_width=True, height=180)
        st.caption(f"{len(df_colab)} linha(s).")

        if st.button("📥 Importar Colaboradores", type="primary", key="btn_colab"):
            def _fn_colab(row):
                nome = _str(row.get("nome"))
                if not nome:
                    return "campo 'nome' obrigatório"
                comissao_pct = _float(row.get("percentual_comissao"))
                if comissao_pct is None:
                    return "campo 'percentual_comissao' obrigatório"
                admissao = _str(row.get("data_admissao")) or str(date.today())
                payload = {
                    "nome":                nome,
                    "cpf":                 _str(row.get("cpf")),
                    "cargo":               _str(row.get("cargo")),
                    "tipo_contrato":       _str(row.get("tipo_contrato")) or "Autônomo",
                    "percentual_comissao": comissao_pct / 100,
                    "telefone":            _str(row.get("telefone")),
                    "email":               _str(row.get("email")),
                    "data_admissao":       admissao,
                }
                resp = api_client.criar_colaborador(payload)
                if resp.status_code not in (200, 201):
                    return resp.json().get("detail", resp.text)
                return None
            _executar(df_colab, _fn_colab, "colaborador(es)")

# ── PROJETOS ──────────────────────────────────────────────────────────────────
with tab_proj:
    st.subheader("Importar Projetos de Homologação")

    COLS_PROJ = [
        {"header": "cliente_nome",    "hint": "Nome exato do cliente",  "req": True,  "w": 28},
        {"header": "uc",              "hint": "3001234567",                            "w": 14},
        {"header": "protocolo",       "hint": "PROT-2026-001",                         "w": 18},
        {"header": "status_projeto",  "hint": "Em Análise",             "w": 16,
         "vals": ["Em Análise", "Aprovado", "Aprovado c/ Obras", "Falta TRT", "Cancelado"]},
        {"header": "status_vistoria", "hint": "Não Solicitado",         "w": 16,
         "vals": ["Não Solicitado", "Solicitado", "Realizado"]},
        {"header": "data_entrada",    "hint": "2026-05-26",             "req": True,  "w": 14},
        {"header": "data_resultado",  "hint": "2026-06-10",                            "w": 14},
        {"header": "kwp",             "hint": "5.5",                                  "w": 10},
        {"header": "inversor",        "hint": "Growatt 5kW",                           "w": 18},
        {"header": "observacoes",     "hint": "Observações livres",                    "w": 28},
    ]

    c1, _ = st.columns([1, 3])
    c1.download_button("⬇️ Template Projetos", _template(COLS_PROJ),
                       "template_projetos.xlsx", XLSX_MIME)

    up = st.file_uploader("Excel preenchido", type="xlsx", key="up_proj")
    if up:
        df_proj = _limpar(pd.read_excel(up, dtype=str))
        st.dataframe(df_proj, use_container_width=True, height=180)
        st.caption(f"{len(df_proj)} linha(s).")

        if st.button("📥 Importar Projetos", type="primary", key="btn_proj"):
            clientes    = api_client.listar_clientes()
            map_cli_id  = {c["nome"].strip().lower(): c["id"] for c in clientes}

            def _fn_proj(row):
                nome_cli = _str(row.get("cliente_nome"))
                if not nome_cli:
                    return "campo 'cliente_nome' obrigatório"
                cliente_id = map_cli_id.get(nome_cli.lower())
                if not cliente_id:
                    return f"cliente '{nome_cli}' não encontrado"
                entrada = _str(row.get("data_entrada")) or str(date.today())
                payload = {
                    "cliente_id":      cliente_id,
                    "uc":              _str(row.get("uc")),
                    "protocolo":       _str(row.get("protocolo")),
                    "status_projeto":  _str(row.get("status_projeto")) or "Em Análise",
                    "status_vistoria": _str(row.get("status_vistoria")) or "Não Solicitado",
                    "data_entrada":    entrada,
                    "data_resultado":  _str(row.get("data_resultado")),
                    "kwp":             _float(row.get("kwp")),
                    "inversor":        _str(row.get("inversor")),
                    "observacoes":     _str(row.get("observacoes")),
                }
                resp = api_client.criar_projeto(payload)
                if resp.status_code not in (200, 201):
                    return resp.json().get("detail", resp.text)
                return None
            _executar(df_proj, _fn_proj, "projeto(s)")
